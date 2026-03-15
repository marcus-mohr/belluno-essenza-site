from __future__ import annotations

import argparse
import json
import sys
import unicodedata
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = PROJECT_ROOT / "data" / "produtos.xlsx"
DEFAULT_JSON = PROJECT_ROOT / "data" / "produtos.json"
DEFAULT_JS = PROJECT_ROOT / "data" / "produtos-data.js"
PRODUCTS_DIR = PROJECT_ROOT / "assets" / "produtos"
BRANDING_FALLBACK = "assets/branding/hero-lifestyle.png"
IMAGE_EXTENSIONS = (".png", ".jpg", ".jpeg", ".webp", ".svg")

EXPECTED_COLUMNS = [
    "ativo",
    "destaque",
    "ordem",
    "slug",
    "nome",
    "subtitulo",
    "categoria",
    "familia_olfativa",
    "preco",
    "preco_promocional",
    "volume",
    "tempo_queima",
    "descricao_curta",
    "descricao_completa",
    "notas_topo",
    "notas_coracao",
    "notas_base",
    "modo_de_uso",
    "composicao",
    "imagem_principal",
    "galeria_1",
    "galeria_2",
    "galeria_3",
    "sku",
    "estoque",
    "selo",
    "meta_title",
    "meta_description",
]

TRUE_VALUES = {"sim", "s", "true", "1", "yes", "y"}
FALSE_VALUES = {"nao", "n", "false", "0", "no"}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Le data/produtos.xlsx e gera data/produtos.json para o site Belluno Essenza."
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Caminho do arquivo Excel de origem.")
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON, help="Caminho do arquivo JSON de saida.")
    parser.add_argument("--js", type=Path, default=DEFAULT_JS, help="Caminho do arquivo JS de fallback local.")
    return parser.parse_args()


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def normalize_ascii(value: object) -> str:
    normalized = unicodedata.normalize("NFD", str(value or "").strip().lower())
    return "".join(char for char in normalized if unicodedata.category(char) != "Mn")


def slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFD", value)
    ascii_text = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    slug = []
    previous_dash = False

    for char in ascii_text.lower():
        if char.isalnum():
            slug.append(char)
            previous_dash = False
        elif not previous_dash:
            slug.append("-")
            previous_dash = True

    return "".join(slug).strip("-")


def as_text(value: object) -> str:
    return str(value or "").strip()


def as_bool(value: object, default: bool = False) -> bool:
    text = normalize_ascii(value)
    if not text:
        return default
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False
    return default


def as_int(value: object) -> int | None:
    text = as_text(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def as_price(value: object) -> float | None:
    text = as_text(value)
    if not text:
        return None

    normalized = text.replace(".", "").replace(",", ".") if "," in text else text

    try:
        return float(Decimal(normalized).quantize(Decimal("0.01")))
    except (InvalidOperation, ValueError):
        return None


def get_headers(worksheet) -> list[str]:
    return [normalize_header(cell.value) for cell in worksheet[1]]


def ensure_expected_columns(headers: list[str]) -> None:
    missing = [column for column in EXPECTED_COLUMNS if column not in headers]
    if missing:
        raise ValueError("A planilha esta sem colunas obrigatorias: " + ", ".join(missing))


def relative_product_path(filename: str) -> str:
    return f"assets/produtos/{filename}"


def candidate_paths(raw_value: str) -> list[Path]:
    value = as_text(raw_value)
    if not value:
        return []

    filename = Path(value).name
    suffix = Path(filename).suffix.lower()

    if suffix:
        return [PRODUCTS_DIR / filename]

    return [PRODUCTS_DIR / f"{filename}{extension}" for extension in IMAGE_EXTENSIONS]


def resolve_explicit_image(row_number: int, field_name: str, raw_value: object, warnings: list[str]) -> str:
    value = as_text(raw_value)
    if not value:
        return ""

    for path in candidate_paths(value):
        if path.exists():
            return relative_product_path(path.name)

    warnings.append(f"Linha {row_number}: arquivo informado em '{field_name}' nao encontrado -> {value}")
    return ""


def resolve_primary_image(row_number: int, slug: str, raw_value: object, warnings: list[str]) -> str:
    explicit_image = resolve_explicit_image(row_number, "imagem_principal", raw_value, warnings)
    if explicit_image:
        return explicit_image

    for path in candidate_paths(slug):
        if path.exists():
            return relative_product_path(path.name)

    warnings.append(
        f"Linha {row_number}: nenhuma imagem principal encontrada. Tentado 'imagem_principal' e fallback por slug '{slug}'."
    )
    return BRANDING_FALLBACK


def build_product(row_number: int, row_data: dict[str, object]) -> tuple[dict[str, object] | None, list[str]]:
    warnings: list[str] = []

    if not as_bool(row_data.get("ativo"), default=True):
        return None, warnings

    name = as_text(row_data.get("nome"))
    if not name:
        warnings.append(f"Linha {row_number}: produto ignorado porque o campo 'nome' esta vazio.")
        return None, warnings

    explicit_slug = as_text(row_data.get("slug"))
    slug = explicit_slug or slugify(name)
    if not slug:
        warnings.append(f"Linha {row_number}: produto ignorado porque nao foi possivel definir o slug.")
        return None, warnings

    normalized_slug = slugify(slug)
    if explicit_slug and normalized_slug != explicit_slug:
        warnings.append(
            f"Linha {row_number}: slug fora do padrao recomendado ('{explicit_slug}'). O valor foi mantido como informado no Excel."
        )

    primary_image = resolve_primary_image(row_number, slug, row_data.get("imagem_principal"), warnings)
    gallery_images = []
    for field_name in ("galeria_1", "galeria_2", "galeria_3"):
        image = resolve_explicit_image(row_number, field_name, row_data.get(field_name), warnings)
        if image:
            gallery_images.append(image)

    images = [primary_image, *gallery_images]

    price = as_price(row_data.get("preco"))
    promotional_price = as_price(row_data.get("preco_promocional"))
    order = as_int(row_data.get("ordem")) or 9999
    stock = as_int(row_data.get("estoque"))

    if promotional_price is not None and price is not None and promotional_price >= price:
        warnings.append(
            f"Linha {row_number}: preco_promocional maior ou igual ao preco principal. O frontend exibira apenas o preco principal."
        )

    short_description = as_text(row_data.get("descricao_curta"))
    full_description = as_text(row_data.get("descricao_completa"))

    product = {
        "ativo": True,
        "destaque": as_bool(row_data.get("destaque")),
        "ordem": order,
        "slug": slug,
        "nome": name,
        "subtitulo": as_text(row_data.get("subtitulo")),
        "categoria": as_text(row_data.get("categoria")),
        "familia_olfativa": as_text(row_data.get("familia_olfativa")),
        "preco": price,
        "preco_promocional": promotional_price,
        "volume": as_text(row_data.get("volume")),
        "tempo_queima": as_text(row_data.get("tempo_queima")),
        "descricao_curta": short_description,
        "descricao_completa": full_description,
        "notas_topo": as_text(row_data.get("notas_topo")),
        "notas_coracao": as_text(row_data.get("notas_coracao")),
        "notas_base": as_text(row_data.get("notas_base")),
        "modo_de_uso": as_text(row_data.get("modo_de_uso")),
        "composicao": as_text(row_data.get("composicao")),
        "imagem_principal": images[0],
        "galeria_1": images[1] if len(images) > 1 else "",
        "galeria_2": images[2] if len(images) > 2 else "",
        "galeria_3": images[3] if len(images) > 3 else "",
        "imagens": images,
        "sku": as_text(row_data.get("sku")),
        "estoque": stock,
        "selo": as_text(row_data.get("selo")),
        "meta_title": as_text(row_data.get("meta_title")) or f"{name} | Belluno Essenza",
        "meta_description": as_text(row_data.get("meta_description")) or short_description or full_description,
        "disponivel": stock is None or stock > 0,
    }

    return product, warnings


def workbook_to_products(xlsx_path: Path) -> tuple[list[dict[str, object]], list[str]]:
    workbook = load_workbook(xlsx_path, data_only=True)
    worksheet = workbook.active

    headers = get_headers(worksheet)
    ensure_expected_columns(headers)

    warnings: list[str] = []
    products: list[dict[str, object]] = []

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        product, row_warnings = build_product(row_number, row_data)
        warnings.extend(row_warnings)
        if product:
            products.append(product)

    products.sort(key=lambda item: (item["ordem"], item["nome"]))
    return products, warnings


def write_json(json_path: Path, products: list[dict[str, object]]) -> None:
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "data/produtos.xlsx",
        "total": len(products),
        "products": products,
    }
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_js(js_path: Path, products: list[dict[str, object]]) -> None:
    payload = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": "data/produtos.xlsx",
        "total": len(products),
        "products": products,
    }
    js_path.parent.mkdir(parents=True, exist_ok=True)
    js_path.write_text(
        "window.BELLUNO_PRODUCTS_DATA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n",
        encoding="utf-8",
    )


def main() -> int:
    args = parse_args()
    xlsx_path = args.xlsx.resolve()
    json_path = args.json.resolve()
    js_path = args.js.resolve()

    if not xlsx_path.exists():
        print(f"Arquivo Excel nao encontrado: {xlsx_path}", file=sys.stderr)
        return 1

    try:
        products, warnings = workbook_to_products(xlsx_path)
        write_json(json_path, products)
        write_js(js_path, products)
    except Exception as error:  # pragma: no cover
        print(f"Erro ao sincronizar produtos: {error}", file=sys.stderr)
        return 1

    print(f"Sincronizacao concluida: {len(products)} produtos ativos gerados em {json_path}")

    if warnings:
        print("\nAvisos:")
        for warning in warnings:
            print(f"- {warning}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
