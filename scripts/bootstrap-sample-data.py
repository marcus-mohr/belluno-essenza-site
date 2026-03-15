from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageColor, ImageDraw, ImageEnhance, ImageFilter, ImageFont, ImageOps


PROJECT_ROOT = Path(__file__).resolve().parents[1]
BRANDING_DIR = PROJECT_ROOT / "assets" / "branding"
PRODUCTS_DIR = PROJECT_ROOT / "assets" / "produtos"
DATA_DIR = PROJECT_ROOT / "data"

LINE_SOURCE = BRANDING_DIR / "linha-de-produtos.png"
LOGO_SOURCE = BRANDING_DIR / "logo-principal.png"
LABEL_SOURCE = BRANDING_DIR / "etiqueta-capim-limao.png"
MOODBOARD_OUTPUT = BRANDING_DIR / "moodboard-belluno.png"

LOGO_BG = "#f3efe8"
GOLD = "#d1bf97"
BROWN = "#5a3d29"
TAUPE = "#8d755c"
GRAPHITE = "#1f1d1d"

COLUMNS = [
    "ativo",
    "destaque",
    "ordem",
    "slug",
    "nome",
    "subtitulo",
    "categoria",
    "familia_olfativa",
    "preco",
    "preco_promocional",
    "volume",
    "tempo_queima",
    "descricao_curta",
    "descricao_completa",
    "notas_topo",
    "notas_coracao",
    "notas_base",
    "modo_de_uso",
    "composicao",
    "imagem_principal",
    "galeria_1",
    "galeria_2",
    "galeria_3",
    "sku",
    "estoque",
    "selo",
    "meta_title",
    "meta_description",
]

PRODUCT_CROPS = {
    "vela-essenza-signature.png": (0, 0, 820, 1024),
    "difusor-essenza-signature.png": (330, 0, 1150, 1024),
    "spray-essenza-signature.png": (760, 120, 1450, 1024),
    "kit-presente-essenza-signature.png": (716, 0, 1536, 1024),
    "colecao-assinatura-belluno.png": (156, 0, 1336, 1024),
    "detalhe-vela-signature.png": (170, 330, 700, 1024),
    "detalhe-difusor-signature.png": (490, 130, 980, 1024),
    "detalhe-spray-signature.png": (770, 240, 1180, 1024),
    "detalhe-caixa-signature.png": (910, 80, 1536, 1024),
}

PRODUCTS = [
    {
        "ativo": "sim",
        "destaque": "sim",
        "ordem": 1,
        "slug": "vela-essenza-signature",
        "nome": "Vela Essenza Signature",
        "subtitulo": "Uma chama macia, com presença elegante e acabamento pensado para aparadores, salas e quartos.",
        "categoria": "Vela Aromática",
        "familia_olfativa": "Âmbar Floral Cremosa",
        "preco": 169.90,
        "preco_promocional": "",
        "volume": "190 g",
        "tempo_queima": "até 40 horas",
        "descricao_curta": "A vela da coleção assinatura Belluno Essenza entrega luz suave, perfume refinado e presença visual sofisticada em qualquer composição de décor.",
        "descricao_completa": "Vela Essenza Signature traduz o estilo da Belluno em uma fragrância confortável, luminosa e elegante. É o tipo de peça que funciona ao mesmo tempo como ritual de pausa, presente delicado e objeto decorativo para uma casa que gosta de receber bem.",
        "notas_topo": "bergamota dourada, folhas macias e acorde luminoso",
        "notas_coracao": "flores brancas, chá claro e toque cremoso",
        "notas_base": "âmbar suave, musk limpo e cedro polido",
        "modo_de_uso": "Na primeira queima, deixe a superfície derreter por completo. Antes de reacender, apare o pavio para cerca de 5 mm.",
        "composicao": "Blend de ceras vegetais, fragrância premium e pavio de algodão.",
        "imagem_principal": "vela-essenza-signature.png",
        "galeria_1": "detalhe-vela-signature.png",
        "galeria_2": "colecao-assinatura-belluno.png",
        "galeria_3": "kit-presente-essenza-signature.png",
        "sku": "BE-VEL-ES-190",
        "estoque": 16,
        "selo": "Best-seller",
        "meta_title": "Vela Essenza Signature | Belluno Essenza",
        "meta_description": "Vela aromática Belluno Essenza com assinatura elegante, luz suave e acabamento premium para a casa.",
    },
    {
        "ativo": "sim",
        "destaque": "sim",
        "ordem": 2,
        "slug": "difusor-essenza-signature",
        "nome": "Difusor Essenza Signature",
        "subtitulo": "Difusão contínua com presença sofisticada e leitura visual impecável sobre aparadores, hall e lavabo.",
        "categoria": "Difusor de Ambiente",
        "familia_olfativa": "Floral Âmbar Elegante",
        "preco": 229.90,
        "preco_promocional": "",
        "volume": "250 ml",
        "tempo_queima": "difusão contínua por até 60 dias",
        "descricao_curta": "Um difusor de presença clássica, com frasco elegante, tampa dourada e varetas pretas que reforçam a assinatura visual da coleção.",
        "descricao_completa": "Difusor Essenza Signature foi pensado para perfumar o ambiente de forma constante, refinada e silenciosa. A composição floral âmbar traz delicadeza na abertura, corpo confortável e um fundo macio que deixa a casa com impressão de cuidado e sofisticação.",
        "notas_topo": "mandarina suave, folhas limpas e chá branco",
        "notas_coracao": "peônia cremosa, muguet e pétalas claras",
        "notas_base": "âmbar confortável, cedro leve e musk branco",
        "modo_de_uso": "Distribua as varetas no frasco e vire-as uma ou duas vezes por semana para renovar a intensidade da fragrância.",
        "composicao": "Base difusora premium, fragrância de ambiente e varetas de fibra preta.",
        "imagem_principal": "difusor-essenza-signature.png",
        "galeria_1": "detalhe-difusor-signature.png",
        "galeria_2": "colecao-assinatura-belluno.png",
        "galeria_3": "kit-presente-essenza-signature.png",
        "sku": "BE-DIF-ES-250",
        "estoque": 13,
        "selo": "Coleção assinatura",
        "meta_title": "Difusor Essenza Signature | Belluno Essenza",
        "meta_description": "Difusor Belluno Essenza com estética clássica, varetas pretas e perfume contínuo para uma casa elegante.",
    },
    {
        "ativo": "sim",
        "destaque": "sim",
        "ordem": 3,
        "slug": "spray-perfumado-essenza-signature",
        "nome": "Spray Perfumado Essenza Signature",
        "subtitulo": "Perfume instantâneo para tecidos, lavabo e os minutos que antecedem uma boa recepção.",
        "categoria": "Spray Perfumado",
        "familia_olfativa": "Âmbar Aromático Luminoso",
        "preco": 139.90,
        "preco_promocional": "",
        "volume": "120 ml",
        "tempo_queima": "aplicação instantânea",
        "descricao_curta": "Uma borrifada elegante, prática e sofisticada para renovar o ambiente com a assinatura olfativa Belluno Essenza.",
        "descricao_completa": "Spray Perfumado Essenza Signature é ideal para quem gosta de ajustar a atmosfera da casa com um gesto rápido e refinado. O frasco esguio, com tampa dourada, acompanha a estética da coleção e oferece perfume imediato para tecidos, lavabo e recepção.",
        "notas_topo": "cítricos polidos, acorde fresco e folhas leves",
        "notas_coracao": "flores claras, chá suave e toque acetinado",
        "notas_base": "âmbar claro, musk confortável e madeira macia",
        "modo_de_uso": "Borrife a cerca de 30 cm de distância sobre o ar ou tecidos delicadamente, evitando superfícies sensíveis e materiais frágeis.",
        "composicao": "Base alcoólica para spray perfumado, fragrância premium e válvula spray.",
        "imagem_principal": "spray-essenza-signature.png",
        "galeria_1": "detalhe-spray-signature.png",
        "galeria_2": "kit-presente-essenza-signature.png",
        "galeria_3": "colecao-assinatura-belluno.png",
        "sku": "BE-SPR-ES-120",
        "estoque": 18,
        "selo": "Novo",
        "meta_title": "Spray Perfumado Essenza Signature | Belluno Essenza",
        "meta_description": "Spray perfumado Belluno Essenza com presença luminosa e acabamento elegante para perfumar a casa em segundos.",
    },
    {
        "ativo": "sim",
        "destaque": "sim",
        "ordem": 4,
        "slug": "kit-presente-essenza-signature",
        "nome": "Kit Presente Essenza Signature",
        "subtitulo": "Um presente completo, com acabamento refinado e atmosfera Belluno pronta para surpreender.",
        "categoria": "Kit Presente",
        "familia_olfativa": "Curadoria Belluno",
        "preco": 319.90,
        "preco_promocional": 289.90,
        "volume": "vela 190 g + spray 120 ml",
        "tempo_queima": "ritual presenteável para uso imediato",
        "descricao_curta": "Uma composição elegante para presentear com delicadeza, reunindo a vela da coleção e o spray perfumado em caixa com laço.",
        "descricao_completa": "Kit Presente Essenza Signature foi pensado para ocasiões em que o gesto importa tanto quanto o produto. A caixa estruturada, o acabamento suave e a combinação entre vela e spray criam uma entrega sofisticada, acolhedora e muito fácil de presentear.",
        "notas_topo": "saída luminosa, fresca e delicadamente cítrica",
        "notas_coracao": "flores claras, chá macio e acorde acetinado",
        "notas_base": "âmbar suave, musk confortável e cedro elegante",
        "modo_de_uso": "Utilize a vela para criar atmosfera e complemente com o spray perfumado para reforçar a sensação de casa pronta e bem recebida.",
        "composicao": "Vela aromática 190 g, spray perfumado 120 ml e caixa presenteável Belluno Essenza.",
        "imagem_principal": "kit-presente-essenza-signature.png",
        "galeria_1": "detalhe-caixa-signature.png",
        "galeria_2": "spray-essenza-signature.png",
        "galeria_3": "vela-essenza-signature.png",
        "sku": "BE-KIT-ES-02",
        "estoque": 8,
        "selo": "Presente Belluno",
        "meta_title": "Kit Presente Essenza Signature | Belluno Essenza",
        "meta_description": "Kit presente Belluno Essenza com vela, spray perfumado e caixa elegante para presentear com sofisticação.",
    },
    {
        "ativo": "sim",
        "destaque": "não",
        "ordem": 5,
        "slug": "colecao-assinatura-belluno",
        "nome": "Coleção Assinatura Belluno",
        "subtitulo": "A linha completa da Belluno Essenza em uma curadoria visual e olfativa para a casa inteira.",
        "categoria": "Kit Presente",
        "familia_olfativa": "Âmbar Floral Amadeirada",
        "preco": 519.90,
        "preco_promocional": 479.90,
        "volume": "vela 190 g + difusor 250 ml + spray 120 ml",
        "tempo_queima": "curadoria completa para presentear ou compor o ambiente",
        "descricao_curta": "A coleção completa reúne os três formatos principais da marca em uma proposta coesa, elegante e altamente presenteável.",
        "descricao_completa": "Coleção Assinatura Belluno foi criada para quem deseja uma experiência completa de ambientação: luz da vela, difusão contínua e perfume imediato. É a seleção mais representativa da identidade Belluno Essenza e traduz bem o posicionamento premium acessível da marca.",
        "notas_topo": "cítricos claros, folhas elegantes e acorde de recepção",
        "notas_coracao": "flores cremosas, chá suave e corpo acetinado",
        "notas_base": "âmbar confortável, musk limpo e madeiras claras",
        "modo_de_uso": "Use o difusor no dia a dia, a vela para criar atmosfera e o spray perfumado para renovar o ambiente antes de receber.",
        "composicao": "Vela aromática 190 g, difusor de ambiente 250 ml, spray perfumado 120 ml e caixa Belluno Essenza.",
        "imagem_principal": "colecao-assinatura-belluno.png",
        "galeria_1": "vela-essenza-signature.png",
        "galeria_2": "difusor-essenza-signature.png",
        "galeria_3": "spray-essenza-signature.png",
        "sku": "BE-COL-ES-03",
        "estoque": 6,
        "selo": "Curadoria completa",
        "meta_title": "Coleção Assinatura Belluno | Belluno Essenza",
        "meta_description": "Linha completa Belluno Essenza com vela, difusor e spray perfumado em uma curadoria elegante para presentear ou compor a casa.",
    },
]


def load_font(path: str, size: int) -> ImageFont.ImageFont:
    try:
        return ImageFont.truetype(path, size)
    except OSError:
        return ImageFont.load_default()


def build_product_images(source: Image.Image) -> None:
    for preview in PRODUCTS_DIR.glob("_preview_*.png"):
        preview.unlink(missing_ok=True)

    for filename, box in PRODUCT_CROPS.items():
        crop = source.crop(box)
        crop = ImageEnhance.Contrast(crop).enhance(1.04)
        crop = ImageEnhance.Color(crop).enhance(1.03)
        crop = ImageEnhance.Brightness(crop).enhance(1.02)
        crop = ImageOps.fit(crop, (1200, 1500), Image.Resampling.LANCZOS, centering=(0.5, 0.45))
        crop.save(PRODUCTS_DIR / filename, quality=95)


def draw_panel_shadow(base: Image.Image, box: tuple[int, int, int, int]) -> Image.Image:
    shadow = Image.new("RGBA", base.size, (0, 0, 0, 0))
    shadow_draw = ImageDraw.Draw(shadow)
    x1, y1, x2, y2 = box
    shadow_draw.rounded_rectangle((x1 + 18, y1 + 22, x2 + 18, y2 + 22), 42, fill=(67, 46, 31, 36))
    return shadow.filter(ImageFilter.GaussianBlur(18))


def draw_text(draw: ImageDraw.ImageDraw, position: tuple[int, int], text: str, font: ImageFont.ImageFont, fill: str) -> None:
    draw.text(position, text, font=font, fill=fill)


def build_moodboard(source: Image.Image) -> None:
    serif_title = load_font(r"C:\Windows\Fonts\georgiab.ttf", 40)
    serif_body = load_font(r"C:\Windows\Fonts\georgia.ttf", 28)
    sans_small = load_font(r"C:\Windows\Fonts\segoeui.ttf", 24)
    sans_caps = load_font(r"C:\Windows\Fonts\seguibli.ttf", 20)

    canvas = Image.new("RGB", (1536, 1536), LOGO_BG)
    panel_boxes = [
        (96, 96, 676, 696),
        (760, 96, 1440, 760),
        (96, 828, 760, 1440),
        (828, 900, 1440, 1440),
    ]

    for box in panel_boxes:
        canvas = Image.alpha_composite(canvas.convert("RGBA"), draw_panel_shadow(canvas, box)).convert("RGB")

    draw = ImageDraw.Draw(canvas)
    fills = ["#fbf8f2", "#f7f2ea", "#f8f3ec", "#f5efe5"]
    for box, fill in zip(panel_boxes, fills):
        draw.rounded_rectangle(box, 42, fill=fill, outline="#dbcba8", width=3)

    logo = Image.open(LOGO_SOURCE).convert("RGBA")
    logo_resized = ImageOps.contain(logo, (430, 430), Image.Resampling.LANCZOS)
    canvas = canvas.convert("RGBA")
    canvas.alpha_composite(logo_resized, (171, 164))
    canvas = canvas.convert("RGB")

    draw = ImageDraw.Draw(canvas)
    draw_text(draw, (164, 522), "COLEÇÃO ASSINATURA", sans_caps, TAUPE)
    draw_text(draw, (164, 560), "Elegância acolhedora em vela, difusor e spray.", serif_body, BROWN)

    top_crop = ImageOps.fit(source.crop((720, 96, 1518, 1024)), (680, 664), Image.Resampling.LANCZOS, centering=(0.58, 0.46))
    top_crop = ImageEnhance.Contrast(top_crop).enhance(1.04)
    canvas.paste(top_crop, (760, 96))

    bottom_crop = ImageOps.fit(source.crop((0, 170, 880, 1024)), (664, 612), Image.Resampling.LANCZOS, centering=(0.42, 0.45))
    bottom_crop = ImageEnhance.Contrast(bottom_crop).enhance(1.03)
    canvas.paste(bottom_crop, (96, 828))

    label = Image.open(LABEL_SOURCE).convert("RGB")
    label_crop = ImageOps.fit(label, (260, 260), Image.Resampling.LANCZOS)
    canvas.paste(label_crop, (866, 952))

    draw = ImageDraw.Draw(canvas)
    draw_text(draw, (1170, 970), "Belluno Essenza", serif_title, BROWN)
    draw_text(draw, (1170, 1026), "Creme quente,\ndourado suave\ne marrom café.", serif_body, TAUPE)
    draw.line((1170, 1126, 1372, 1126), fill=GOLD, width=2)
    draw_text(draw, (1170, 1152), "PALETA", sans_caps, TAUPE)

    swatches = [LOGO_BG, "#eadcc3", "#d8c494", BROWN, GRAPHITE]
    for index, color in enumerate(swatches):
        x = 1170 + index * 52
        draw.rounded_rectangle((x, 1196, x + 40, 1236), 10, fill=ImageColor.getrgb(color), outline="#d4c29b")

    draw_text(draw, (866, 1288), "Branding correto,\npaleta suave e atmosfera editorial.", serif_body, TAUPE)
    draw_text(draw, (866, 1364), "Composição criada a partir da\nlinha real da Belluno Essenza.", sans_small, TAUPE)

    canvas.save(MOODBOARD_OUTPUT, quality=95)


def build_workbook() -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Produtos"
    worksheet.append(COLUMNS)

    header_fill = PatternFill(fill_type="solid", fgColor="4B3527")
    header_font = Font(color="F7F2EA", bold=True)
    side = Side(style="thin", color="D6C6A4")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=side, right=side, top=side, bottom=side)

    for product in PRODUCTS:
        worksheet.append([product[column] for column in COLUMNS])

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=side, right=side, top=side, bottom=side)

    widths = {
        "A": 10, "B": 12, "C": 8, "D": 34, "E": 34, "F": 54, "G": 24, "H": 28,
        "I": 12, "J": 18, "K": 26, "L": 34, "M": 58, "N": 72, "O": 38, "P": 38,
        "Q": 38, "R": 52, "S": 44, "T": 34, "U": 32, "V": 32, "W": 32, "X": 18,
        "Y": 12, "Z": 20, "AA": 48, "AB": 62,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width

    worksheet.freeze_panes = "A2"
    return workbook


def main() -> int:
    BRANDING_DIR.mkdir(parents=True, exist_ok=True)
    PRODUCTS_DIR.mkdir(parents=True, exist_ok=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if not LINE_SOURCE.exists():
        raise FileNotFoundError(f"Arquivo de referência não encontrado: {LINE_SOURCE}")
    if not LOGO_SOURCE.exists():
        raise FileNotFoundError(f"Logo local não encontrado: {LOGO_SOURCE}")
    if not LABEL_SOURCE.exists():
        raise FileNotFoundError(f"Etiqueta local não encontrada: {LABEL_SOURCE}")

    source = Image.open(LINE_SOURCE).convert("RGB")
    build_product_images(source)
    build_moodboard(source)

    workbook = build_workbook()
    workbook.save(DATA_DIR / "produtos.xlsx")

    print("Produtos exemplo, moodboard Belluno e planilha foram regenerados com a linha real da marca.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
